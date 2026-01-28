from flask import Flask, request, jsonify
from flask_cors import CORS
import os
from openpyxl import load_workbook

app = Flask(__name__)
CORS(app)

# 模拟已有的“系统资源库”
# 匹配规则依然是：资源名称 + 频道
DATABASE_RESOURCES = {
    ("白月光", "电影频道"),
    ("狂飙", "电视剧频道"),
    ("流浪地球", "科幻频道")
}

@app.route('/api/audit/upload', methods=['POST'])
def upload_audit():
    if 'file' not in request.files:
        return jsonify({"error": "未找到文件"}), 400
    
    file = request.files['file']
    
    try:
        wb = load_workbook(file)
        sheet = wb.active
        
        results = []
        # max_col=3：读取 A(业务线), B(资源名称), C(频道)
        for row in sheet.iter_rows(min_row=1, max_col=3, values_only=True):
            biz_line, res_name, channel = row
            
            if res_name is None: continue 
            
            # 数据清洗
            biz_line = str(biz_line).strip() if biz_line else "未知业务线"
            res_name = str(res_name).strip()
            channel = str(channel).strip() if channel else "未知频道"
            
            # 查重逻辑
            is_dup = (res_name, channel) in DATABASE_RESOURCES
            
            results.append({
                "biz_line": biz_line,
                "name": res_name,
                "channel": channel,
                "status": "重复资源" if is_dup else "独家资源", # 对齐需求文案
                "is_duplicate": is_dup
            })
        
        return jsonify(results)
    except Exception as e:
        return jsonify({"error": f"解析失败: {str(e)}"}), 500

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 8080))
    app.run(host='0.0.0.0', port=port)